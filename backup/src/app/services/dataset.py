import hashlib
import json
import re
from collections import defaultdict
from collections.abc import Callable
from datetime import date
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, func, insert, or_, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.db.tables import table
from app.models import DocumentVersion, ProcessingJob
from app.schemas.dataset import (
    DatasetBuildCreate,
    DatasetCellUpdate,
    DatasetFieldCreate,
    DatasetRowCreate,
    DatasetVersionClone,
)
from app.schemas.workflow import TaskAccepted
from app.services.audit import AuditService
from app.services.storage import LocalStorage


def _has_dataset_value(cell: dict[str, Any] | None) -> bool:
    if not cell or cell.get("is_missing"):
        return False
    value_columns = (
        "raw_value",
        "normalized_value",
        "ml_value",
        "value_text",
        "value_number",
        "value_boolean",
        "value_date",
        "value_json",
        "range_min",
        "range_max",
        "mean_value",
        "standard_deviation",
        "significance_marker",
    )
    for column in value_columns:
        value = cell.get(column)
        if isinstance(value, str):
            if value.strip():
                return True
        elif isinstance(value, (dict, list, tuple, set)):
            if value:
                return True
        elif value is not None:
            return True
    return False


def _freeze_issues(
    fields: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    cells: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    cells_by_position = {(cell["row_id"], cell["field_id"]): cell for cell in cells}
    issues: list[dict[str, Any]] = []
    for row in rows:
        for field in fields:
            cell = cells_by_position.get((row["id"], field["id"]))
            context = {
                "row_id": str(row["id"]),
                "row_no": row["row_no"],
                "row_key": row["row_key"],
                "field_id": str(field["id"]),
                "field_key": field["field_key"],
                "display_name": field["display_name"],
            }
            if field["is_required"] and not _has_dataset_value(cell):
                issues.append({"code": "required_value_missing", **context})
            if cell and cell.get("review_status") == "doubtful":
                issues.append(
                    {
                        "code": "doubtful_value",
                        "cell_id": str(cell["id"]),
                        **context,
                    }
                )
    return issues


def _freeze_snapshot(
    fields: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    cells: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
) -> dict[str, Any]:
    field_key = {field["id"]: field["field_key"] for field in fields}
    row_key = {row["id"]: row["row_key"] for row in rows}
    cell_position = {
        cell["id"]: (row_key[cell["row_id"]], field_key[cell["field_id"]]) for cell in cells
    }
    field_columns = (
        "field_key",
        "display_name",
        "data_type",
        "semantic_role",
        "unit_id",
        "position",
        "is_required",
        "is_hidden",
        "validation_rules",
        "display_config",
        "metadata",
    )
    row_columns = (
        "row_no",
        "row_key",
        "source_document_id",
        "source_document_version_id",
        "source_sample_key",
        "review_status",
        "metadata",
    )
    cell_columns = (
        "source_extraction_record_id",
        "raw_value",
        "raw_unit_text",
        "normalized_value",
        "ml_value",
        "value_text",
        "value_number",
        "value_boolean",
        "value_date",
        "value_json",
        "range_min",
        "range_max",
        "mean_value",
        "standard_deviation",
        "significance_marker",
        "unit_id",
        "value_source",
        "review_status",
        "confidence",
        "is_missing",
        "is_image_estimate",
        "is_manually_modified",
        "notes",
        "metadata",
    )
    evidence_columns = (
        "extraction_evidence_id",
        "document_version_id",
        "page_id",
        "block_id",
        "table_cell_id",
        "figure_id",
        "evidence_text",
        "bbox",
        "is_primary",
    )
    return {
        "fields": [{column: field[column] for column in field_columns} for field in fields],
        "rows": [{column: row[column] for column in row_columns} for row in rows],
        "cells": [
            {
                "row_key": row_key[cell["row_id"]],
                "field_key": field_key[cell["field_id"]],
                **{column: cell[column] for column in cell_columns},
            }
            for cell in sorted(
                cells,
                key=lambda item: (
                    row_key[item["row_id"]],
                    field_key[item["field_id"]],
                ),
            )
        ],
        "evidence": [
            {
                "row_key": cell_position[item["dataset_cell_id"]][0],
                "field_key": cell_position[item["dataset_cell_id"]][1],
                **{column: item[column] for column in evidence_columns},
            }
            for item in sorted(
                evidence,
                key=lambda item: (
                    cell_position[item["dataset_cell_id"]],
                    not item["is_primary"],
                    item["evidence_text"],
                    str(item["page_id"]),
                ),
            )
        ],
    }


EXPORT_MISSING = "——"
BASE_EXPORT_COLUMNS = (
    "来源文件名",
    "篇名",
    "出版时间",
    "出版物名称",
    "品种/处理组",
    "地点/材料",
    "时间点",
    "证据状态",
)

_EXPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EXPORT_HEADER_FILL = PatternFill("solid", fgColor="305496")


def _style_export_header(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = _EXPORT_HEADER_FONT
        cell.fill = _EXPORT_HEADER_FILL


def _sample_key_value(sample_key: Any, key: str) -> str | None:
    if not sample_key:
        return None
    match = re.search(rf"{key}=([^:]+)", str(sample_key))
    return match.group(1) if match else None


def _publication_time(document: dict[str, Any] | None) -> str:
    if not document:
        return EXPORT_MISSING
    if document.get("publication_date"):
        return str(document["publication_date"])
    if document.get("publication_year"):
        return str(document["publication_year"])
    return EXPORT_MISSING


def _metadata_value(metadata: Any, *keys: str) -> str | None:
    if isinstance(metadata, dict):
        for key in keys:
            value = metadata.get(key)
            if value:
                return str(value)
    return None


def _base_row_values(
    row: dict[str, Any], document: dict[str, Any] | None, filename: str | None
) -> list[Any]:
    row_metadata = row.get("metadata") or {}
    document_metadata = (document or {}).get("metadata")
    treatment = (
        _metadata_value(row_metadata, "treatment", "variety")
        or _sample_key_value(row.get("source_sample_key"), "treatment")
        or _metadata_value(document_metadata, "variety", "treatment")
        or EXPORT_MISSING
    )
    location = (
        _metadata_value(document_metadata, "location", "material")
        or _metadata_value(row_metadata, "location", "material")
        or EXPORT_MISSING
    )
    timepoint = (
        _metadata_value(row_metadata, "timepoint")
        or _sample_key_value(row.get("source_sample_key"), "time")
        or EXPORT_MISSING
    )
    return [
        filename or EXPORT_MISSING,
        (document or {}).get("title") or EXPORT_MISSING,
        _publication_time(document),
        (document or {}).get("publication_name") or EXPORT_MISSING,
        treatment,
        location,
        timepoint,
        row.get("review_status") or EXPORT_MISSING,
    ]


def _raw_cell_value(cell: dict[str, Any] | None) -> Any:
    if not cell:
        return EXPORT_MISSING
    if cell.get("raw_value") is not None:
        return cell["raw_value"]
    if cell.get("value_number") is not None:
        return cell["value_number"]
    return cell.get("value_text") or EXPORT_MISSING


def _ml_cell_value(cell: dict[str, Any] | None) -> Any:
    if not cell:
        return ""
    if cell.get("value_number") is not None:
        return cell["value_number"]
    if cell.get("value_boolean") is not None:
        return cell["value_boolean"]
    if cell.get("value_date") is not None:
        return str(cell["value_date"])
    if cell.get("value_text"):
        return cell["value_text"]
    for column in ("normalized_value", "ml_value"):
        value = cell.get(column)
        if isinstance(value, dict) and value:
            if "value" in value:
                return value["value"]
            return json.dumps(value, ensure_ascii=False, default=str)
    return ""


def _json_cell(value: Any) -> str:
    if value in (None, "", {}, []):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


class DatasetService:
    def __init__(self, db: Session, storage: LocalStorage) -> None:
        self.db = db
        self.storage = storage

    def create_from_extraction(
        self,
        project_id: UUID,
        payload: DatasetBuildCreate,
        actor_id: UUID | None,
    ) -> TaskAccepted:
        extraction_runs = table(self.db, "extraction_runs")
        extraction = (
            self.db.execute(
                select(extraction_runs).where(
                    extraction_runs.c.id == payload.extraction_run_id,
                    extraction_runs.c.project_id == project_id,
                )
            )
            .mappings()
            .one_or_none()
        )
        if not extraction:
            raise AppError(
                code="extraction_run_not_found", message="抽取任务不存在", status_code=404
            )
        if extraction["status"] != "completed":
            raise AppError(
                code="extraction_not_completed", message="抽取完成后才能生成数据集", status_code=409
            )
        datasets = table(self.db, "datasets")
        versions = table(self.db, "dataset_versions")
        dataset_id = self.db.execute(
            insert(datasets)
            .values(
                project_id=project_id,
                name=payload.name,
                description=payload.description,
                purpose="research",
                status="active",
                settings={},
                created_by=actor_id,
            )
            .returning(datasets.c.id)
        ).scalar_one()
        version_id = self.db.execute(
            insert(versions)
            .values(
                dataset_id=dataset_id,
                version_no=1,
                field_schema_id=extraction["field_schema_id"],
                source_extraction_run_id=payload.extraction_run_id,
                status="draft",
                metadata={"include_review_statuses": payload.include_review_statuses},
                created_by=actor_id,
            )
            .returning(versions.c.id)
        ).scalar_one()
        job = ProcessingJob(
            project_id=project_id,
            job_type="build_dataset",
            status="queued",
            progress_percent=0,
            current_stage="waiting",
            idempotency_key=f"build_dataset:{version_id}",
            requested_config={"dataset_version_id": str(version_id)},
            result_summary={},
            requested_by=actor_id,
        )
        self.db.add(job)
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset",
            entity_id=dataset_id,
            action="create",
            after={"name": payload.name, "version_id": str(version_id)},
        )
        self.db.commit()
        return TaskAccepted(resource_id=version_id, job_id=job.id)

    def _version_context(self, project_id: UUID, version_id: UUID) -> tuple[dict, dict]:
        datasets = table(self.db, "datasets")
        versions = table(self.db, "dataset_versions")
        row = (
            self.db.execute(
                select(datasets, versions)
                .join(versions, versions.c.dataset_id == datasets.c.id)
                .where(
                    versions.c.id == version_id,
                    datasets.c.project_id == project_id,
                    datasets.c.deleted_at.is_(None),
                )
            )
            .mappings()
            .one_or_none()
        )
        if not row:
            raise AppError(
                code="dataset_version_not_found", message="数据集版本不存在", status_code=404
            )
        dataset_data = {column.name: row[column] for column in datasets.c}
        version_data = {column.name: row[column] for column in versions.c}
        return dataset_data, version_data

    def list_datasets(self, project_id: UUID) -> list[dict]:
        datasets = table(self.db, "datasets")
        versions = table(self.db, "dataset_versions")
        latest = (
            select(versions.c.dataset_id, func.max(versions.c.version_no).label("version_no"))
            .group_by(versions.c.dataset_id)
            .subquery()
        )
        rows = (
            self.db.execute(
                select(
                    datasets,
                    versions.c.id.label("latest_version_id"),
                    versions.c.version_no.label("latest_version_no"),
                    versions.c.status.label("latest_version_status"),
                    versions.c.row_count,
                    versions.c.field_count,
                )
                .join(latest, latest.c.dataset_id == datasets.c.id)
                .join(
                    versions,
                    (versions.c.dataset_id == latest.c.dataset_id)
                    & (versions.c.version_no == latest.c.version_no),
                )
                .where(datasets.c.project_id == project_id, datasets.c.deleted_at.is_(None))
                .order_by(datasets.c.created_at.desc())
            )
            .mappings()
            .all()
        )
        return [dict(row) for row in rows]

    def list_versions(self, project_id: UUID, dataset_id: UUID) -> list[dict]:
        datasets = table(self.db, "datasets")
        versions = table(self.db, "dataset_versions")
        if not self.db.scalar(
            select(datasets.c.id).where(
                datasets.c.id == dataset_id,
                datasets.c.project_id == project_id,
                datasets.c.deleted_at.is_(None),
            )
        ):
            raise AppError(code="dataset_not_found", message="数据集不存在", status_code=404)
        rows = self.db.execute(
            select(versions)
            .where(versions.c.dataset_id == dataset_id)
            .order_by(versions.c.version_no.desc())
        ).mappings()
        return [dict(row) for row in rows]

    def get_version(self, project_id: UUID, version_id: UUID, offset: int, limit: int) -> dict:
        dataset_data, version_data = self._version_context(project_id, version_id)
        fields = table(self.db, "dataset_fields")
        rows = table(self.db, "dataset_rows")
        cells = table(self.db, "dataset_cells")
        evidence = table(self.db, "dataset_cell_evidence")
        document_versions = table(self.db, "document_versions")
        documents = table(self.db, "documents")
        pages = table(self.db, "document_pages")
        field_rows = [
            dict(row)
            for row in self.db.execute(
                select(fields)
                .where(fields.c.dataset_version_id == version_id)
                .order_by(fields.c.position)
            ).mappings()
        ]
        row_rows = [
            dict(row)
            for row in self.db.execute(
                select(rows)
                .where(rows.c.dataset_version_id == version_id, rows.c.is_deleted.is_(False))
                .order_by(rows.c.row_no)
                .offset(offset)
                .limit(limit)
            ).mappings()
        ]
        row_ids = [row["id"] for row in row_rows]
        cell_rows = (
            self.db.execute(select(cells).where(cells.c.row_id.in_(row_ids))).mappings().all()
            if row_ids
            else []
        )
        cell_ids = [cell["id"] for cell in cell_rows]
        evidence_rows = (
            self.db.execute(
                select(
                    evidence,
                    documents.c.id.label("document_id"),
                    documents.c.title.label("document_title"),
                    pages.c.page_no,
                    pages.c.width.label("page_width"),
                    pages.c.height.label("page_height"),
                )
                .join(
                    document_versions,
                    document_versions.c.id == evidence.c.document_version_id,
                )
                .join(documents, documents.c.id == document_versions.c.document_id)
                .join(pages, pages.c.id == evidence.c.page_id)
                .where(evidence.c.dataset_cell_id.in_(cell_ids))
                .order_by(
                    evidence.c.dataset_cell_id,
                    evidence.c.is_primary.desc(),
                    evidence.c.created_at,
                )
            )
            .mappings()
            .all()
            if cell_ids
            else []
        )
        evidence_by_cell: dict[UUID, list[dict]] = defaultdict(list)
        for item in evidence_rows:
            evidence_by_cell[item["dataset_cell_id"]].append(dict(item))
        by_row: dict[UUID, dict[str, dict]] = defaultdict(dict)
        field_key = {field["id"]: field["field_key"] for field in field_rows}
        for cell in cell_rows:
            cell_data = dict(cell)
            cell_data["evidence"] = evidence_by_cell.get(cell["id"], [])
            by_row[cell["row_id"]][field_key[cell["field_id"]]] = cell_data
        for row in row_rows:
            row["cells"] = by_row.get(row["id"], {})
        return {
            "dataset": dataset_data,
            "version": version_data,
            "fields": field_rows,
            "rows": row_rows,
            "offset": offset,
            "limit": limit,
        }

    def add_field(
        self, project_id: UUID, version_id: UUID, payload: DatasetFieldCreate, actor_id: UUID | None
    ) -> dict:
        _, version = self._version_context(project_id, version_id)
        if version["status"] != "draft":
            raise AppError(
                code="dataset_immutable", message="只有草稿版本可增加字段", status_code=409
            )
        fields = table(self.db, "dataset_fields")
        position = (
            self.db.scalar(
                select(func.max(fields.c.position)).where(fields.c.dataset_version_id == version_id)
            )
            or -1
        ) + 1
        row = (
            self.db.execute(
                insert(fields)
                .values(
                    dataset_version_id=version_id,
                    field_key=payload.field_key,
                    display_name=payload.display_name,
                    data_type=payload.data_type,
                    semantic_role=payload.semantic_role,
                    unit_id=payload.unit_id,
                    position=position,
                    is_required=payload.is_required,
                    validation_rules={},
                    display_config={},
                    metadata={},
                )
                .returning(fields)
            )
            .mappings()
            .one()
        )
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_field",
            entity_id=row["id"],
            action="create",
            after=dict(row),
        )
        self.db.commit()
        return dict(row)

    def add_row(
        self, project_id: UUID, version_id: UUID, payload: DatasetRowCreate, actor_id: UUID | None
    ) -> dict:
        _, version = self._version_context(project_id, version_id)
        if version["status"] != "draft":
            raise AppError(
                code="dataset_immutable", message="只有草稿版本可增加行", status_code=409
            )
        rows = table(self.db, "dataset_rows")
        row_no = (
            self.db.scalar(
                select(func.max(rows.c.row_no)).where(rows.c.dataset_version_id == version_id)
            )
            or 0
        ) + 1
        row = (
            self.db.execute(
                insert(rows)
                .values(
                    dataset_version_id=version_id,
                    row_no=row_no,
                    row_key=payload.row_key,
                    source_document_id=payload.source_document_id,
                    source_document_version_id=payload.source_document_version_id,
                    review_status="pending",
                    is_deleted=False,
                    metadata=payload.metadata,
                    created_by=actor_id,
                )
                .returning(rows)
            )
            .mappings()
            .one()
        )
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_row",
            entity_id=row["id"],
            action="create",
            after=dict(row),
        )
        self.db.commit()
        return dict(row)

    def update_cell(
        self,
        project_id: UUID,
        version_id: UUID,
        row_id: UUID,
        field_id: UUID,
        payload: DatasetCellUpdate,
        actor_id: UUID | None,
    ) -> dict:
        _, version = self._version_context(project_id, version_id)
        if version["status"] != "draft":
            raise AppError(code="dataset_immutable", message="冻结版本不可修改", status_code=409)
        rows = table(self.db, "dataset_rows")
        fields = table(self.db, "dataset_fields")
        cells = table(self.db, "dataset_cells")
        if not self.db.scalar(
            select(rows.c.id).where(rows.c.id == row_id, rows.c.dataset_version_id == version_id)
        ):
            raise AppError(code="dataset_row_not_found", message="数据行不存在", status_code=404)
        if not self.db.scalar(
            select(fields.c.id).where(
                fields.c.id == field_id, fields.c.dataset_version_id == version_id
            )
        ):
            raise AppError(
                code="dataset_field_not_found", message="数据字段不存在", status_code=404
            )
        before = (
            self.db.execute(
                select(cells).where(cells.c.row_id == row_id, cells.c.field_id == field_id)
            )
            .mappings()
            .one_or_none()
        )
        values = payload.model_dump(exclude_unset=True)
        if "value_date" in values and isinstance(values["value_date"], str):
            values["value_date"] = date.fromisoformat(values["value_date"])
        values["is_manually_modified"] = True
        values["modified_by"] = actor_id
        statement = (
            pg_insert(cells)
            .values(row_id=row_id, field_id=field_id, value_source="manual", **values)
            .on_conflict_do_update(index_elements=[cells.c.row_id, cells.c.field_id], set_=values)
            .returning(cells)
        )
        row = self.db.execute(statement).mappings().one()
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_cell",
            entity_id=row["id"],
            action="update" if before else "create",
            before=dict(before) if before else None,
            after=dict(row),
        )
        self.db.commit()
        return dict(row)

    def delete_row(
        self, project_id: UUID, version_id: UUID, row_id: UUID, actor_id: UUID | None
    ) -> None:
        _, version = self._version_context(project_id, version_id)
        if version["status"] != "draft":
            raise AppError(code="dataset_immutable", message="冻结版本不可修改", status_code=409)
        rows = table(self.db, "dataset_rows")
        result = self.db.execute(
            update(rows)
            .where(rows.c.id == row_id, rows.c.dataset_version_id == version_id)
            .values(is_deleted=True, review_status="deleted")
        )
        if not getattr(result, "rowcount", 0):
            raise AppError(code="dataset_row_not_found", message="数据行不存在", status_code=404)
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_row",
            entity_id=row_id,
            action="soft_delete",
        )
        self.db.commit()

    def freeze(self, project_id: UUID, version_id: UUID, actor_id: UUID | None) -> dict:
        _, version = self._version_context(project_id, version_id)
        if version["status"] != "draft":
            raise AppError(
                code="dataset_not_freezable", message="只有草稿版本可冻结", status_code=409
            )
        fields = table(self.db, "dataset_fields")
        rows = table(self.db, "dataset_rows")
        cells = table(self.db, "dataset_cells")
        evidence = table(self.db, "dataset_cell_evidence")
        field_rows = [
            dict(row)
            for row in self.db.execute(
                select(fields)
                .where(fields.c.dataset_version_id == version_id)
                .order_by(fields.c.position, fields.c.field_key)
            ).mappings()
        ]
        row_rows = [
            dict(row)
            for row in self.db.execute(
                select(rows)
                .where(rows.c.dataset_version_id == version_id, rows.c.is_deleted.is_(False))
                .order_by(rows.c.row_no, rows.c.row_key)
            ).mappings()
        ]
        row_ids = [row["id"] for row in row_rows]
        cell_rows = (
            [
                dict(row)
                for row in self.db.execute(
                    select(cells)
                    .where(cells.c.row_id.in_(row_ids))
                    .order_by(cells.c.row_id, cells.c.field_id)
                ).mappings()
            ]
            if row_ids
            else []
        )
        cell_ids = [cell["id"] for cell in cell_rows]
        evidence_rows = (
            [
                dict(row)
                for row in self.db.execute(
                    select(evidence)
                    .where(evidence.c.dataset_cell_id.in_(cell_ids))
                    .order_by(
                        evidence.c.dataset_cell_id,
                        evidence.c.is_primary.desc(),
                        evidence.c.created_at,
                    )
                ).mappings()
            ]
            if cell_ids
            else []
        )
        issues = _freeze_issues(field_rows, row_rows, cell_rows)
        if issues:
            raise AppError(
                code="dataset_freeze_validation_failed",
                message="数据集存在必填缺失或疑似值，无法冻结",
                status_code=409,
                details={"issue_count": len(issues), "issues": issues},
            )
        snapshot = _freeze_snapshot(field_rows, row_rows, cell_rows, evidence_rows)
        content_hash = hashlib.sha256(
            json.dumps(snapshot, ensure_ascii=False, default=str, sort_keys=True).encode()
        ).hexdigest()
        versions = table(self.db, "dataset_versions")
        row_count = (
            self.db.scalar(
                select(func.count())
                .select_from(rows)
                .where(rows.c.dataset_version_id == version_id, rows.c.is_deleted.is_(False))
            )
            or 0
        )
        field_count = (
            self.db.scalar(
                select(func.count())
                .select_from(fields)
                .where(fields.c.dataset_version_id == version_id)
            )
            or 0
        )
        updated = (
            self.db.execute(
                update(versions)
                .where(versions.c.id == version_id, versions.c.status == "draft")
                .values(
                    status="frozen",
                    content_sha256=content_hash,
                    row_count=row_count,
                    field_count=field_count,
                    frozen_by=actor_id,
                    frozen_at=func.now(),
                )
                .returning(versions)
            )
            .mappings()
            .one()
        )
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_version",
            entity_id=version_id,
            action="freeze",
            after={"sha256": content_hash},
        )
        self.db.commit()
        return dict(updated)

    def clone_version(
        self,
        project_id: UUID,
        version_id: UUID,
        payload: DatasetVersionClone,
        actor_id: UUID | None,
    ) -> dict:
        dataset, source = self._version_context(project_id, version_id)
        versions = table(self.db, "dataset_versions")
        fields = table(self.db, "dataset_fields")
        rows = table(self.db, "dataset_rows")
        cells = table(self.db, "dataset_cells")
        evidence = table(self.db, "dataset_cell_evidence")
        conversions = table(self.db, "conversion_records")
        next_version = (
            self.db.scalar(
                select(func.max(versions.c.version_no)).where(
                    versions.c.dataset_id == source["dataset_id"]
                )
            )
            or 0
        ) + 1
        new_version_id = self.db.execute(
            insert(versions)
            .values(
                dataset_id=source["dataset_id"],
                version_no=next_version,
                parent_version_id=version_id,
                field_schema_id=source["field_schema_id"],
                source_extraction_run_id=source["source_extraction_run_id"],
                status="draft",
                change_summary=payload.change_summary,
                row_count=source["row_count"],
                field_count=source["field_count"],
                metadata={
                    **(source["metadata"] or {}),
                    "cloned_from_version_id": str(version_id),
                },
                created_by=actor_id,
            )
            .returning(versions.c.id)
        ).scalar_one()
        field_map: dict[UUID, UUID] = {}
        for item in self.db.execute(
            select(fields)
            .where(fields.c.dataset_version_id == version_id)
            .order_by(fields.c.position)
        ).mappings():
            field_map[item["id"]] = self.db.execute(
                insert(fields)
                .values(
                    dataset_version_id=new_version_id,
                    source_field_id=item["source_field_id"],
                    field_key=item["field_key"],
                    display_name=item["display_name"],
                    data_type=item["data_type"],
                    semantic_role=item["semantic_role"],
                    unit_id=item["unit_id"],
                    position=item["position"],
                    is_required=item["is_required"],
                    is_hidden=item["is_hidden"],
                    validation_rules=item["validation_rules"],
                    display_config=item["display_config"],
                    metadata=item["metadata"],
                )
                .returning(fields.c.id)
            ).scalar_one()
        row_map: dict[UUID, UUID] = {}
        for item in self.db.execute(
            select(rows).where(rows.c.dataset_version_id == version_id).order_by(rows.c.row_no)
        ).mappings():
            row_map[item["id"]] = self.db.execute(
                insert(rows)
                .values(
                    dataset_version_id=new_version_id,
                    row_no=item["row_no"],
                    row_key=item["row_key"],
                    source_document_id=item["source_document_id"],
                    source_document_version_id=item["source_document_version_id"],
                    source_sample_key=item["source_sample_key"],
                    review_status=item["review_status"],
                    is_deleted=item["is_deleted"],
                    metadata=item["metadata"],
                    created_by=actor_id,
                )
                .returning(rows.c.id)
            ).scalar_one()
        cell_map: dict[UUID, UUID] = {}
        source_cells = self.db.execute(
            select(cells).where(cells.c.row_id.in_(list(row_map)))
        ).mappings()
        excluded = {"id", "row_id", "field_id", "created_at", "updated_at"}
        for item in source_cells:
            values = {key: value for key, value in item.items() if key not in excluded}
            values.update(row_id=row_map[item["row_id"]], field_id=field_map[item["field_id"]])
            cell_map[item["id"]] = self.db.execute(
                insert(cells).values(**values).returning(cells.c.id)
            ).scalar_one()
        if cell_map:
            excluded_conversion = {"id", "dataset_cell_id", "created_at"}
            for item in self.db.execute(
                select(conversions).where(conversions.c.dataset_cell_id.in_(list(cell_map)))
            ).mappings():
                values = {
                    key: value for key, value in item.items() if key not in excluded_conversion
                }
                values["dataset_cell_id"] = cell_map[item["dataset_cell_id"]]
                self.db.execute(insert(conversions).values(**values))
        if cell_map:
            excluded_evidence = {"id", "dataset_cell_id", "created_at"}
            for item in self.db.execute(
                select(evidence).where(evidence.c.dataset_cell_id.in_(list(cell_map)))
            ).mappings():
                values = {key: value for key, value in item.items() if key not in excluded_evidence}
                values["dataset_cell_id"] = cell_map[item["dataset_cell_id"]]
                self.db.execute(insert(evidence).values(**values))
        AuditService(self.db).record(
            project_id=project_id,
            actor_id=actor_id,
            entity_type="dataset_version",
            entity_id=new_version_id,
            action="clone",
            before={"version_id": str(version_id), "version_no": source["version_no"]},
            after={"version_id": str(new_version_id), "version_no": next_version},
            reason=payload.change_summary,
        )
        self.db.commit()
        return {
            "dataset_id": dataset["id"],
            "version_id": new_version_id,
            "version_no": next_version,
            "parent_version_id": version_id,
            "status": "draft",
        }

    def build(self, version_id: UUID, progress: Callable[[float, str], None]) -> dict[str, Any]:
        datasets = table(self.db, "datasets")
        versions = table(self.db, "dataset_versions")
        dataset_fields = table(self.db, "dataset_fields")
        dataset_rows = table(self.db, "dataset_rows")
        dataset_cells = table(self.db, "dataset_cells")
        cell_evidence = table(self.db, "dataset_cell_evidence")
        source_fields = table(self.db, "field_definitions")
        records = table(self.db, "extraction_records")
        evidence = table(self.db, "extraction_evidence")
        conversions = table(self.db, "conversion_records")
        version = (
            self.db.execute(
                select(versions, datasets.c.project_id)
                .join(datasets, datasets.c.id == versions.c.dataset_id)
                .where(versions.c.id == version_id)
            )
            .mappings()
            .one_or_none()
        )
        if not version:
            raise AppError(
                code="dataset_version_not_found", message="数据集版本不存在", status_code=404
            )
        self.db.execute(delete(dataset_rows).where(dataset_rows.c.dataset_version_id == version_id))
        self.db.execute(
            delete(dataset_fields).where(dataset_fields.c.dataset_version_id == version_id)
        )
        field_rows = [
            dict(row)
            for row in self.db.execute(
                select(source_fields)
                .where(source_fields.c.field_schema_id == version["field_schema_id"])
                .order_by(source_fields.c.position)
            ).mappings()
        ]
        field_map: dict[UUID, UUID] = {}
        for source in field_rows:
            field_id = self.db.execute(
                insert(dataset_fields)
                .values(
                    dataset_version_id=version_id,
                    source_field_id=source["id"],
                    field_key=source["field_key"],
                    display_name=source["display_name"],
                    data_type=source["data_type"],
                    semantic_role=source["semantic_role"],
                    unit_id=source["preferred_unit_id"],
                    position=source["position"],
                    is_required=source["is_required"],
                    validation_rules=source["validation_rules"],
                    display_config=source["display_config"],
                    metadata={},
                )
                .returning(dataset_fields.c.id)
            ).scalar_one()
            field_map[source["id"]] = field_id
        progress(20, "created_fields")

        statuses = (version["metadata"] or {}).get(
            "include_review_statuses", ["pending", "confirmed", "modified"]
        )
        record_rows = [
            dict(row)
            for row in self.db.execute(
                select(records)
                .where(
                    records.c.extraction_run_id == version["source_extraction_run_id"],
                    records.c.review_status.in_(statuses),
                )
                .order_by(records.c.created_at)
            ).mappings()
        ]
        grouped: dict[str, list[dict]] = defaultdict(list)
        for record in record_rows:
            grouped[record["group_key"] or record["sample_key"]].append(record)
        for row_no, (row_key, group_records) in enumerate(grouped.items(), start=1):
            document_version_id = group_records[0]["document_version_id"]
            document_id = self.db.scalar(
                select(DocumentVersion.document_id).where(DocumentVersion.id == document_version_id)
            )
            row_id = self.db.execute(
                insert(dataset_rows)
                .values(
                    dataset_version_id=version_id,
                    row_no=row_no,
                    row_key=row_key,
                    source_document_id=document_id,
                    source_document_version_id=document_version_id,
                    source_sample_key=group_records[0]["sample_key"],
                    review_status="pending",
                    is_deleted=False,
                    metadata={"record_count": len(group_records)},
                )
                .returning(dataset_rows.c.id)
            ).scalar_one()
            used_fields: set[UUID] = set()
            for record in group_records:
                target_field = field_map.get(record["field_definition_id"])
                if not target_field or target_field in used_fields:
                    continue
                used_fields.add(target_field)
                parsed = record["parsed_value"] or {}
                cell_id = self.db.execute(
                    insert(dataset_cells)
                    .values(
                        row_id=row_id,
                        field_id=target_field,
                        source_extraction_record_id=record["id"],
                        raw_value=record["raw_value"],
                        raw_unit_text=record["raw_unit_text"],
                        normalized_value=record["normalized_value"],
                        ml_value=record["ml_value"],
                        value_number=record["numeric_value"],
                        range_min=record["range_min"],
                        range_max=record["range_max"],
                        mean_value=record["mean_value"],
                        standard_deviation=record["standard_deviation"],
                        significance_marker=record["significance_marker"],
                        unit_id=record["normalized_unit_id"],
                        value_source="extracted",
                        review_status=record["review_status"],
                        confidence=record["confidence"],
                        is_missing=record["is_missing"],
                        is_image_estimate=record["is_image_estimate"],
                        is_manually_modified=False,
                        metadata={"parsed_value": parsed},
                    )
                    .returning(dataset_cells.c.id)
                ).scalar_one()
                self.db.execute(
                    update(conversions)
                    .where(conversions.c.extraction_record_id == record["id"])
                    .values(dataset_cell_id=cell_id)
                )
                evidence_rows = (
                    self.db.execute(
                        select(evidence).where(evidence.c.extraction_record_id == record["id"])
                    )
                    .mappings()
                    .all()
                )
                for item in evidence_rows:
                    self.db.execute(
                        insert(cell_evidence).values(
                            dataset_cell_id=cell_id,
                            extraction_evidence_id=item["id"],
                            document_version_id=item["document_version_id"],
                            page_id=item["page_id"],
                            block_id=item["block_id"],
                            table_cell_id=item["table_cell_id"],
                            figure_id=item["figure_id"],
                            evidence_text=item["evidence_text"],
                            bbox=item["bbox"],
                            is_primary=item["is_primary"],
                        )
                    )
            if row_no % 50 == 0:
                progress(20 + 75 * row_no / max(len(grouped), 1), "building_rows")
                self.db.commit()
        self.db.execute(
            update(versions)
            .where(versions.c.id == version_id)
            .values(row_count=len(grouped), field_count=len(field_rows))
        )
        self.db.commit()
        return {
            "dataset_version_id": str(version_id),
            "row_count": len(grouped),
            "field_count": len(field_rows),
        }

    def _export_document_context(
        self, rows: list[dict[str, Any]]
    ) -> tuple[dict[UUID, dict[str, Any]], dict[UUID, str]]:
        document_ids = {
            row["source_document_id"] for row in rows if row.get("source_document_id")
        }
        version_ids = {
            row["source_document_version_id"]
            for row in rows
            if row.get("source_document_version_id")
        }
        documents_by_id: dict[UUID, dict[str, Any]] = {}
        if document_ids:
            documents = table(self.db, "documents")
            for document in self.db.execute(
                select(documents).where(documents.c.id.in_(document_ids))
            ).mappings():
                documents_by_id[document["id"]] = dict(document)
        filename_by_version: dict[UUID, str] = {}
        if version_ids:
            document_versions = table(self.db, "document_versions")
            stored_files = table(self.db, "stored_files")
            for version_id, original_name in self.db.execute(
                select(document_versions.c.id, stored_files.c.original_name)
                .join(stored_files, stored_files.c.id == document_versions.c.source_file_id)
                .where(document_versions.c.id.in_(version_ids))
            ):
                if original_name:
                    filename_by_version[version_id] = original_name
        return documents_by_id, filename_by_version

    def export_xlsx(self, project_id: UUID, version_id: UUID) -> Path:
        dataset_data, version_data = self._version_context(project_id, version_id)
        content = self.get_version(
            project_id, version_id, 0, max(int(version_data["row_count"] or 0), 1_000_000)
        )
        fields = content["fields"]
        rows = content["rows"]
        documents_by_id, filename_by_version = self._export_document_context(rows)
        field_headers = [field["display_name"] for field in fields]

        workbook = Workbook()
        main_sheet = workbook.active
        main_sheet.title = "dataset_main"
        main_sheet.append([*BASE_EXPORT_COLUMNS, *field_headers])
        _style_export_header(main_sheet)

        ml_sheet = workbook.create_sheet("dataset_ml")
        ml_sheet.append([*BASE_EXPORT_COLUMNS, *field_headers])
        _style_export_header(ml_sheet)

        for row in rows:
            document = documents_by_id.get(row.get("source_document_id"))
            filename = filename_by_version.get(row.get("source_document_version_id"))
            base_values = _base_row_values(row, document, filename)
            main_values = list(base_values)
            ml_values = list(base_values)
            for field in fields:
                cell = row["cells"].get(field["field_key"])
                main_values.append(_raw_cell_value(cell))
                ml_values.append(_ml_cell_value(cell))
            main_sheet.append(main_values)
            ml_sheet.append(ml_values)

        self._write_token_dictionary(workbook.create_sheet("token_dictionary"), project_id)
        self._write_traceability(workbook.create_sheet("traceability"), version_id)
        self._write_conversion_records(workbook.create_sheet("conversion_records"), project_id)
        self._write_audit_log(workbook.create_sheet("audit_log"), project_id)

        output_dir = self.storage.path_for_key(f"exports/{project_id}")
        output_dir.mkdir(parents=True, exist_ok=True)
        output = output_dir / f"dataset-{dataset_data['id']}-v{version_data['version_no']}.xlsx"
        workbook.save(output)
        return output

    def _write_token_dictionary(self, sheet: Worksheet, project_id: UUID) -> None:
        sheet.append(["标准名", "类别", "别名", "首选单位", "状态"])
        _style_export_header(sheet)
        terms = table(self.db, "terms")
        categories = table(self.db, "term_categories")
        aliases = table(self.db, "term_aliases")
        units = table(self.db, "units")
        term_rows = (
            self.db.execute(
                select(
                    terms.c.id,
                    terms.c.canonical_name,
                    terms.c.status,
                    categories.c.name.label("category_name"),
                    units.c.symbol.label("unit_symbol"),
                )
                .select_from(terms)
                .join(categories, categories.c.id == terms.c.category_id, isouter=True)
                .join(units, units.c.id == terms.c.preferred_unit_id, isouter=True)
                .where(
                    terms.c.project_id == project_id,
                    terms.c.deleted_at.is_(None),
                    or_(terms.c.status == "confirmed", terms.c.is_selected.is_(True)),
                )
                .order_by(terms.c.canonical_name)
            )
            .mappings()
            .all()
        )
        if not term_rows:
            return
        alias_map: dict[UUID, list[str]] = defaultdict(list)
        for term_id, alias_text in self.db.execute(
            select(aliases.c.term_id, aliases.c.alias_text)
            .where(aliases.c.term_id.in_([row["id"] for row in term_rows]))
            .order_by(aliases.c.alias_text)
        ):
            if alias_text:
                alias_map[term_id].append(str(alias_text))
        for row in term_rows:
            sheet.append(
                [
                    row["canonical_name"],
                    row["category_name"] or EXPORT_MISSING,
                    "、".join(alias_map.get(row["id"], [])) or EXPORT_MISSING,
                    row["unit_symbol"] or EXPORT_MISSING,
                    row["status"] or EXPORT_MISSING,
                ]
            )

    def _write_traceability(self, sheet: Worksheet, version_id: UUID) -> None:
        sheet.append(["row_key", "field_key", "raw_value", "page_no", "evidence_text"])
        _style_export_header(sheet)
        evidence_table = table(self.db, "dataset_cell_evidence")
        cells = table(self.db, "dataset_cells")
        rows_table = table(self.db, "dataset_rows")
        fields_table = table(self.db, "dataset_fields")
        pages = table(self.db, "document_pages")
        trace_rows = self.db.execute(
            select(
                rows_table.c.row_key,
                fields_table.c.field_key,
                cells.c.raw_value,
                pages.c.page_no,
                evidence_table.c.evidence_text,
            )
            .join(cells, cells.c.row_id == rows_table.c.id)
            .join(fields_table, fields_table.c.id == cells.c.field_id)
            .join(evidence_table, evidence_table.c.dataset_cell_id == cells.c.id)
            .join(pages, pages.c.id == evidence_table.c.page_id)
            .where(rows_table.c.dataset_version_id == version_id)
        ).all()
        for row in trace_rows:
            sheet.append(list(row))

    def _write_conversion_records(self, sheet: Worksheet, project_id: UUID) -> None:
        sheet.append(["原值", "原单位", "标准值", "标准单位", "规则/公式", "状态", "时间"])
        _style_export_header(sheet)
        conversions = table(self.db, "conversion_records")
        units = table(self.db, "units")
        source_units = units.alias("source_units")
        target_units = units.alias("target_units")
        rows = self.db.execute(
            select(
                conversions,
                source_units.c.symbol.label("source_symbol"),
                target_units.c.symbol.label("target_symbol"),
            )
            .select_from(conversions)
            .join(source_units, source_units.c.id == conversions.c.source_unit_id, isouter=True)
            .join(target_units, target_units.c.id == conversions.c.target_unit_id, isouter=True)
            .where(conversions.c.project_id == project_id)
            .order_by(conversions.c.created_at)
        ).mappings()
        for row in rows:
            record = dict(row)
            sheet.append(
                [
                    _json_cell(record.get("source_value")),
                    record.get("source_symbol") or record.get("source_unit_text") or EXPORT_MISSING,
                    _json_cell(record.get("target_value")),
                    record.get("target_symbol") or EXPORT_MISSING,
                    record.get("formula_used") or EXPORT_MISSING,
                    record.get("status") or EXPORT_MISSING,
                    str(record.get("created_at")) if record.get("created_at") else EXPORT_MISSING,
                ]
            )

    def _write_audit_log(self, sheet: Worksheet, project_id: UUID) -> None:
        sheet.append(["时间", "操作", "对象类型", "对象ID", "执行人", "摘要"])
        _style_export_header(sheet)
        logs = table(self.db, "audit_logs")
        users = table(self.db, "app_users")
        rows = self.db.execute(
            select(logs, users.c.display_name.label("actor_name"))
            .select_from(logs)
            .join(users, users.c.id == logs.c.actor_id, isouter=True)
            .where(logs.c.project_id == project_id)
            .order_by(logs.c.created_at.desc())
            .limit(5000)
        ).mappings()
        for row in rows:
            record = dict(row)
            summary = record.get("reason") or _json_cell(record.get("after_value")) or ""
            actor = record.get("actor_name") or (
                str(record["actor_id"]) if record.get("actor_id") else EXPORT_MISSING
            )
            sheet.append(
                [
                    str(record.get("created_at")) if record.get("created_at") else EXPORT_MISSING,
                    record.get("action") or EXPORT_MISSING,
                    record.get("entity_type") or EXPORT_MISSING,
                    str(record["entity_id"]) if record.get("entity_id") else EXPORT_MISSING,
                    actor,
                    summary,
                ]
            )
