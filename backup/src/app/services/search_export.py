import json
from collections import defaultdict
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.db.tables import table
from app.services.storage import LocalStorage

MISSING = "——"
STRUCTURED_CONTENT_LIMIT = 2000

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


def _style_header(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _authors_text(authors: Any) -> str:
    if not authors:
        return MISSING
    names: list[str] = []
    for author in authors:
        if isinstance(author, dict):
            names.append(str(author.get("name") or author.get("display_name") or "").strip())
        else:
            names.append(str(author).strip())
    return "、".join(name for name in names if name) or MISSING


def _publication_time(document: dict[str, Any]) -> str:
    if document.get("publication_date"):
        return str(document["publication_date"])
    if document.get("publication_year"):
        return str(document["publication_year"])
    return MISSING


def _metadata_value(metadata: Any, *keys: str) -> str:
    if isinstance(metadata, dict):
        for key in keys:
            value = metadata.get(key)
            if value:
                return str(value)
    return MISSING


def _matched_terms_text(matched_terms: Any) -> str:
    if not isinstance(matched_terms, list):
        return ""
    matched: list[str] = []
    everything: list[str] = []
    for item in matched_terms:
        if isinstance(item, dict):
            term = item.get("term")
            if not term:
                continue
            everything.append(str(term))
            if item.get("matched"):
                matched.append(str(term))
        else:
            everything.append(str(item))
    selected = matched or everything
    return "、".join(dict.fromkeys(selected))


def _combined_evidence(row: dict[str, Any]) -> str:
    parts = [row.get("previous_context"), row.get("matched_context"), row.get("next_context")]
    joined = " ".join(str(part) for part in parts if part)
    return joined or MISSING


def _structured_content(payload: Any) -> str:
    if payload in (None, "", {}, []):
        return ""
    if not isinstance(payload, (str, bytes)):
        payload = json.dumps(payload, ensure_ascii=False, default=str)
    text = payload.decode() if isinstance(payload, bytes) else str(payload)
    return text[:STRUCTURED_CONTENT_LIMIT]


class SearchExportService:
    def __init__(self, db: Session, storage: LocalStorage) -> None:
        self.db = db
        self.storage = storage

    def export_xlsx(self, project_id: UUID, run_id: UUID) -> Path:
        runs = table(self.db, "search_runs")
        run = (
            self.db.execute(
                select(runs).where(runs.c.id == run_id, runs.c.project_id == project_id)
            )
            .mappings()
            .one_or_none()
        )
        if not run:
            raise AppError(code="search_run_not_found", message="检索任务不存在", status_code=404)

        results = table(self.db, "search_results")
        version_ids = list(
            self.db.scalars(
                select(results.c.document_version_id)
                .where(results.c.search_run_id == run_id)
                .distinct()
            ).all()
        )

        workbook = Workbook()
        self._write_documents(workbook.active, run_id)
        self._write_search_results(workbook.create_sheet("search_results"), run_id)
        self._write_figure_table_results(
            workbook.create_sheet("figure_table_results"), version_ids
        )
        self._write_token_candidates(workbook.create_sheet("token_candidates"), project_id)

        output_dir = self.storage.path_for_key(f"exports/{project_id}")
        output_dir.mkdir(parents=True, exist_ok=True)
        output = output_dir / f"search-run-{run_id}.xlsx"
        workbook.save(output)
        return output

    def _write_documents(self, sheet: Worksheet, run_id: UUID) -> None:
        sheet.title = "documents"
        sheet.append(
            ["标题", "作者", "出版时间", "出版物", "品种/处理组", "地点/材料", "文件名"]
        )
        _style_header(sheet)
        results = table(self.db, "search_results")
        document_versions = table(self.db, "document_versions")
        documents = table(self.db, "documents")
        stored_files = table(self.db, "stored_files")
        rows = self.db.execute(
            select(
                documents.c.title,
                documents.c.authors,
                documents.c.publication_date,
                documents.c.publication_year,
                documents.c.publication_name,
                documents.c.metadata,
                stored_files.c.original_name,
                func.min(results.c.result_no).label("order_no"),
            )
            .select_from(results)
            .join(document_versions, document_versions.c.id == results.c.document_version_id)
            .join(documents, documents.c.id == document_versions.c.document_id)
            .join(
                stored_files,
                stored_files.c.id == document_versions.c.source_file_id,
                isouter=True,
            )
            .where(results.c.search_run_id == run_id)
            .group_by(
                documents.c.id,
                documents.c.title,
                documents.c.authors,
                documents.c.publication_date,
                documents.c.publication_year,
                documents.c.publication_name,
                documents.c.metadata,
                stored_files.c.original_name,
            )
            .order_by(func.min(results.c.result_no))
        ).mappings()
        for row in rows:
            document = dict(row)
            sheet.append(
                [
                    document.get("title") or MISSING,
                    _authors_text(document.get("authors")),
                    _publication_time(document),
                    document.get("publication_name") or MISSING,
                    _metadata_value(document.get("metadata"), "variety", "treatment"),
                    _metadata_value(document.get("metadata"), "location", "material"),
                    document.get("original_name") or MISSING,
                ]
            )

    def _write_search_results(self, sheet: Worksheet, run_id: UUID) -> None:
        sheet.append(
            [
                "序号",
                "文献ID",
                "标题",
                "检索词",
                "证据类型",
                "页码",
                "上一句",
                "命中句",
                "下一句",
                "合并证据内容",
                "审核状态",
            ]
        )
        _style_header(sheet)
        results = table(self.db, "search_results")
        pages = table(self.db, "document_pages")
        document_versions = table(self.db, "document_versions")
        documents = table(self.db, "documents")
        rows = self.db.execute(
            select(
                results,
                pages.c.page_no,
                documents.c.id.label("document_id"),
                documents.c.title.label("document_title"),
            )
            .select_from(results)
            .join(pages, pages.c.id == results.c.page_id, isouter=True)
            .join(document_versions, document_versions.c.id == results.c.document_version_id)
            .join(documents, documents.c.id == document_versions.c.document_id)
            .where(results.c.search_run_id == run_id)
            .order_by(results.c.result_no)
        ).mappings()
        for row in rows:
            result = dict(row)
            sheet.append(
                [
                    result.get("result_no"),
                    str(result.get("document_id")) if result.get("document_id") else MISSING,
                    result.get("document_title") or MISSING,
                    _matched_terms_text(result.get("matched_terms")),
                    result.get("evidence_type") or MISSING,
                    result.get("page_no"),
                    result.get("previous_context") or "",
                    result.get("matched_context") or "",
                    result.get("next_context") or "",
                    _combined_evidence(result),
                    result.get("review_status") or MISSING,
                ]
            )

    def _write_figure_table_results(self, sheet: Worksheet, version_ids: list[UUID]) -> None:
        sheet.append(
            ["表号/图号", "表题/图题", "图注", "结构化内容", "页码", "所属文献"]
        )
        _style_header(sheet)
        if not version_ids:
            return
        pages = table(self.db, "document_pages")
        document_versions = table(self.db, "document_versions")
        documents = table(self.db, "documents")
        tables = table(self.db, "document_tables")
        figures = table(self.db, "document_figures")

        table_rows = self.db.execute(
            select(
                tables.c.table_no,
                tables.c.title,
                tables.c.caption,
                tables.c.structured_data,
                pages.c.page_no,
                documents.c.title.label("document_title"),
            )
            .select_from(tables)
            .join(pages, pages.c.id == tables.c.page_id, isouter=True)
            .join(document_versions, document_versions.c.id == tables.c.document_version_id)
            .join(documents, documents.c.id == document_versions.c.document_id)
            .where(tables.c.document_version_id.in_(version_ids))
            .order_by(tables.c.document_version_id, tables.c.table_no)
        ).mappings()
        for row in table_rows:
            item = dict(row)
            sheet.append(
                [
                    item.get("table_no") or MISSING,
                    item.get("title") or MISSING,
                    item.get("caption") or MISSING,
                    _structured_content(item.get("structured_data")),
                    item.get("page_no"),
                    item.get("document_title") or MISSING,
                ]
            )

        figure_rows = self.db.execute(
            select(
                figures.c.figure_no,
                figures.c.title,
                figures.c.caption,
                figures.c.axis_metadata,
                figures.c.legend_metadata,
                figures.c.extracted_labels,
                figures.c.semantic_summary,
                pages.c.page_no,
                documents.c.title.label("document_title"),
            )
            .select_from(figures)
            .join(pages, pages.c.id == figures.c.page_id, isouter=True)
            .join(document_versions, document_versions.c.id == figures.c.document_version_id)
            .join(documents, documents.c.id == document_versions.c.document_id)
            .where(figures.c.document_version_id.in_(version_ids))
            .order_by(figures.c.document_version_id, figures.c.figure_no)
        ).mappings()
        for row in figure_rows:
            item = dict(row)
            structured = {
                "axis_metadata": item.get("axis_metadata"),
                "legend_metadata": item.get("legend_metadata"),
                "extracted_labels": item.get("extracted_labels"),
                "semantic_summary": item.get("semantic_summary"),
            }
            sheet.append(
                [
                    item.get("figure_no") or MISSING,
                    item.get("title") or MISSING,
                    item.get("caption") or MISSING,
                    _structured_content(structured),
                    item.get("page_no"),
                    item.get("document_title") or MISSING,
                ]
            )

    def _write_token_candidates(self, sheet: Worksheet, project_id: UUID) -> None:
        sheet.append(
            ["候选词", "类别", "出现次数", "出现文献数", "别名建议", "示例证据", "人工选择状态"]
        )
        _style_header(sheet)
        terms = table(self.db, "terms")
        categories = table(self.db, "term_categories")
        aliases = table(self.db, "term_aliases")
        occurrences = table(self.db, "term_occurrences")
        term_rows = self.db.execute(
            select(
                terms.c.id,
                terms.c.canonical_name,
                terms.c.status,
                terms.c.is_selected,
                categories.c.name.label("category_name"),
            )
            .select_from(terms)
            .join(categories, categories.c.id == terms.c.category_id, isouter=True)
            .where(terms.c.project_id == project_id, terms.c.deleted_at.is_(None))
            .order_by(terms.c.canonical_name)
        ).mappings().all()
        if not term_rows:
            return
        term_ids = [row["id"] for row in term_rows]

        alias_map: dict[UUID, list[str]] = defaultdict(list)
        for term_id, alias_text in self.db.execute(
            select(aliases.c.term_id, aliases.c.alias_text)
            .where(aliases.c.term_id.in_(term_ids))
            .order_by(aliases.c.alias_text)
        ):
            if alias_text:
                alias_map[term_id].append(str(alias_text))

        occurrence_stats: dict[UUID, tuple[int, int]] = {}
        for term_id, total, documents_count in self.db.execute(
            select(
                occurrences.c.term_id,
                func.coalesce(func.sum(occurrences.c.occurrence_count), 0),
                func.count(func.distinct(occurrences.c.document_version_id)),
            )
            .where(occurrences.c.term_id.in_(term_ids))
            .group_by(occurrences.c.term_id)
        ):
            occurrence_stats[term_id] = (int(total or 0), int(documents_count or 0))

        example_map: dict[UUID, str] = {}
        for term_id, context_text in self.db.execute(
            select(occurrences.c.term_id, occurrences.c.context_text).where(
                occurrences.c.term_id.in_(term_ids)
            )
        ):
            if term_id not in example_map and context_text:
                example_map[term_id] = str(context_text)

        for row in term_rows:
            term_id = row["id"]
            total, documents_count = occurrence_stats.get(term_id, (0, 0))
            selection = "已选择" if row["is_selected"] else (row["status"] or MISSING)
            sheet.append(
                [
                    row["canonical_name"],
                    row["category_name"] or MISSING,
                    total,
                    documents_count,
                    "、".join(alias_map.get(term_id, [])) or MISSING,
                    example_map.get(term_id, MISSING),
                    selection,
                ]
            )
