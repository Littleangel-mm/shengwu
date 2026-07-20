from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import (
    JSON,
    Boolean,
    Column,
    Date,
    DateTime,
    Integer,
    MetaData,
    Numeric,
    String,
    Table,
    Text,
    Uuid,
    create_engine,
)
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.core.config import Settings
from app.core.errors import AppError
from app.services import dataset as dataset_module
from app.services import search_export as search_export_module
from app.services.dataset import DatasetService
from app.services.search_export import SearchExportService
from app.services.storage import LocalStorage


def _now() -> datetime:
    return datetime.now(UTC)


def _build_tables(engine) -> dict[str, Table]:
    metadata = MetaData()
    definitions: dict[str, list[Column]] = {
        "app_users": [
            Column("id", Uuid, primary_key=True),
            Column("display_name", String(200)),
        ],
        "units": [
            Column("id", Uuid, primary_key=True),
            Column("symbol", String(80)),
        ],
        "documents": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("title", Text),
            Column("authors", JSON),
            Column("publication_date", Date),
            Column("publication_year", Integer),
            Column("publication_name", Text),
            Column("metadata", JSON),
        ],
        "document_versions": [
            Column("id", Uuid, primary_key=True),
            Column("document_id", Uuid),
            Column("source_file_id", Uuid),
        ],
        "stored_files": [
            Column("id", Uuid, primary_key=True),
            Column("original_name", Text),
        ],
        "document_pages": [
            Column("id", Uuid, primary_key=True),
            Column("page_no", Integer),
            Column("width", Numeric),
            Column("height", Numeric),
        ],
        "document_tables": [
            Column("id", Uuid, primary_key=True),
            Column("document_version_id", Uuid),
            Column("page_id", Uuid),
            Column("table_no", String(100)),
            Column("title", Text),
            Column("caption", Text),
            Column("structured_data", JSON),
        ],
        "document_figures": [
            Column("id", Uuid, primary_key=True),
            Column("document_version_id", Uuid),
            Column("page_id", Uuid),
            Column("figure_no", String(100)),
            Column("title", Text),
            Column("caption", Text),
            Column("axis_metadata", JSON),
            Column("legend_metadata", JSON),
            Column("extracted_labels", JSON),
            Column("semantic_summary", Text),
        ],
        "search_runs": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
        ],
        "search_results": [
            Column("id", Uuid, primary_key=True),
            Column("search_run_id", Uuid),
            Column("result_no", Integer),
            Column("document_version_id", Uuid),
            Column("page_id", Uuid),
            Column("evidence_type", String(32)),
            Column("previous_context", Text),
            Column("matched_context", Text),
            Column("next_context", Text),
            Column("matched_terms", JSON),
            Column("review_status", String(32)),
        ],
        "datasets": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("name", String(240)),
            Column("deleted_at", DateTime(timezone=True)),
        ],
        "dataset_versions": [
            Column("id", Uuid, primary_key=True),
            Column("dataset_id", Uuid),
            Column("version_no", Integer),
            Column("status", String(32)),
            Column("row_count", Integer),
            Column("field_count", Integer),
            Column("metadata", JSON),
        ],
        "dataset_fields": [
            Column("id", Uuid, primary_key=True),
            Column("dataset_version_id", Uuid),
            Column("field_key", String(160)),
            Column("display_name", String(240)),
            Column("position", Integer),
        ],
        "dataset_rows": [
            Column("id", Uuid, primary_key=True),
            Column("dataset_version_id", Uuid),
            Column("row_no", Integer),
            Column("row_key", String(255)),
            Column("source_document_id", Uuid),
            Column("source_document_version_id", Uuid),
            Column("source_sample_key", String(255)),
            Column("review_status", String(32)),
            Column("is_deleted", Boolean),
            Column("metadata", JSON),
        ],
        "dataset_cells": [
            Column("id", Uuid, primary_key=True),
            Column("row_id", Uuid),
            Column("field_id", Uuid),
            Column("raw_value", Text),
            Column("normalized_value", JSON),
            Column("ml_value", JSON),
            Column("value_text", Text),
            Column("value_number", Numeric),
            Column("value_boolean", Boolean),
            Column("value_date", Date),
            Column("value_json", JSON),
        ],
        "dataset_cell_evidence": [
            Column("id", Uuid, primary_key=True),
            Column("dataset_cell_id", Uuid),
            Column("document_version_id", Uuid),
            Column("page_id", Uuid),
            Column("block_id", Uuid),
            Column("table_cell_id", Uuid),
            Column("figure_id", Uuid),
            Column("extraction_evidence_id", Uuid),
            Column("evidence_text", Text),
            Column("bbox", JSON),
            Column("is_primary", Boolean),
            Column("created_at", DateTime(timezone=True)),
        ],
        "terms": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("category_id", Uuid),
            Column("canonical_name", Text),
            Column("status", String(32)),
            Column("is_selected", Boolean),
            Column("preferred_unit_id", Uuid),
            Column("deleted_at", DateTime(timezone=True)),
        ],
        "term_categories": [
            Column("id", Uuid, primary_key=True),
            Column("name", String(200)),
        ],
        "term_aliases": [
            Column("id", Uuid, primary_key=True),
            Column("term_id", Uuid),
            Column("alias_text", Text),
        ],
        "term_occurrences": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("term_id", Uuid),
            Column("document_version_id", Uuid),
            Column("context_text", Text),
            Column("occurrence_count", Integer),
        ],
        "conversion_records": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("source_value", JSON),
            Column("source_unit_text", String(160)),
            Column("source_unit_id", Uuid),
            Column("target_value", JSON),
            Column("target_unit_id", Uuid),
            Column("formula_used", Text),
            Column("status", String(32)),
            Column("created_at", DateTime(timezone=True)),
        ],
        "audit_logs": [
            Column("id", Uuid, primary_key=True),
            Column("project_id", Uuid),
            Column("actor_id", Uuid),
            Column("action", String(100)),
            Column("entity_type", String(100)),
            Column("entity_id", Uuid),
            Column("reason", Text),
            Column("after_value", JSON),
            Column("created_at", DateTime(timezone=True)),
        ],
    }
    tables = {name: Table(name, metadata, *columns) for name, columns in definitions.items()}
    metadata.create_all(engine)
    return tables


@pytest.fixture
def export_env(monkeypatch, tmp_path: Path):
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    tables = _build_tables(engine)
    monkeypatch.setattr(dataset_module, "table", lambda _db, name: tables[name])
    monkeypatch.setattr(search_export_module, "table", lambda _db, name: tables[name])
    session = Session(engine)
    storage = LocalStorage(Settings(storage_root=tmp_path / "storage"))
    try:
        yield session, tables, storage
    finally:
        session.close()


def _insert_common_document(session: Session, tables: dict[str, Table], project_id):
    document_id = uuid4()
    version_id = uuid4()
    file_id = uuid4()
    page_id = uuid4()
    session.execute(
        tables["stored_files"].insert().values(id=file_id, original_name="paperA.pdf")
    )
    session.execute(
        tables["documents"].insert().values(
            id=document_id,
            project_id=project_id,
            title="论文A",
            authors=["李雷", "韩梅梅"],
            publication_date=None,
            publication_year=2023,
            publication_name="生物学报",
            metadata={"variety": "品种X", "location": "北京"},
        )
    )
    session.execute(
        tables["document_versions"].insert().values(
            id=version_id, document_id=document_id, source_file_id=file_id
        )
    )
    session.execute(
        tables["document_pages"].insert().values(
            id=page_id, page_no=3, width=595, height=842
        )
    )
    return document_id, version_id, page_id


def test_dataset_export_has_six_sheets_with_expected_headers(export_env) -> None:
    session, tables, storage = export_env
    project_id = uuid4()
    document_id, version_id, page_id = _insert_common_document(session, tables, project_id)

    user_id = uuid4()
    unit_id = uuid4()
    dataset_id = uuid4()
    dataset_version_id = uuid4()
    field_id = uuid4()
    row_id = uuid4()
    cell_id = uuid4()
    category_id = uuid4()
    term_id = uuid4()

    session.execute(tables["app_users"].insert().values(id=user_id, display_name="张三"))
    session.execute(tables["units"].insert().values(id=unit_id, symbol="%"))
    session.execute(
        tables["datasets"].insert().values(
            id=dataset_id, project_id=project_id, name="数据集A", deleted_at=None
        )
    )
    session.execute(
        tables["dataset_versions"].insert().values(
            id=dataset_version_id,
            dataset_id=dataset_id,
            version_no=1,
            status="frozen",
            row_count=1,
            field_count=1,
            metadata={},
        )
    )
    session.execute(
        tables["dataset_fields"].insert().values(
            id=field_id,
            dataset_version_id=dataset_version_id,
            field_key="temperature",
            display_name="温度",
            position=0,
        )
    )
    session.execute(
        tables["dataset_rows"].insert().values(
            id=row_id,
            dataset_version_id=dataset_version_id,
            row_no=1,
            row_key="sample-1",
            source_document_id=document_id,
            source_document_version_id=version_id,
            source_sample_key="doc:treatment=A:time=48h",
            review_status="confirmed",
            is_deleted=False,
            metadata={},
        )
    )
    session.execute(
        tables["dataset_cells"].insert().values(
            id=cell_id,
            row_id=row_id,
            field_id=field_id,
            raw_value="25",
            normalized_value={"value": 25},
            ml_value={"value": 0.25},
            value_text=None,
            value_number=25,
            value_boolean=None,
            value_date=None,
            value_json=None,
        )
    )
    session.execute(
        tables["dataset_cell_evidence"].insert().values(
            id=uuid4(),
            dataset_cell_id=cell_id,
            document_version_id=version_id,
            page_id=page_id,
            is_primary=True,
            evidence_text="温度25℃",
            created_at=_now(),
        )
    )
    session.execute(
        tables["term_categories"].insert().values(id=category_id, name="工艺参数")
    )
    session.execute(
        tables["terms"].insert().values(
            id=term_id,
            project_id=project_id,
            category_id=category_id,
            canonical_name="温度",
            status="confirmed",
            is_selected=True,
            preferred_unit_id=unit_id,
            deleted_at=None,
        )
    )
    session.execute(
        tables["term_aliases"].insert().values(id=uuid4(), term_id=term_id, alias_text="temp")
    )
    session.execute(
        tables["conversion_records"].insert().values(
            id=uuid4(),
            project_id=project_id,
            source_value={"value": 25},
            source_unit_text="C",
            source_unit_id=unit_id,
            target_value={"value": 298},
            target_unit_id=unit_id,
            formula_used="x+273.15",
            status="confirmed",
            created_at=_now(),
        )
    )
    session.execute(
        tables["audit_logs"].insert().values(
            id=uuid4(),
            project_id=project_id,
            actor_id=user_id,
            action="update",
            entity_type="dataset_cell",
            entity_id=cell_id,
            reason="人工修改",
            after_value={"value_number": 25},
            created_at=_now(),
        )
    )
    session.commit()

    path = DatasetService(session, storage).export_xlsx(project_id, dataset_version_id)
    workbook = load_workbook(BytesIO(Path(path).read_bytes()))

    assert workbook.sheetnames == [
        "dataset_main",
        "dataset_ml",
        "token_dictionary",
        "traceability",
        "conversion_records",
        "audit_log",
    ]

    base = [
        "来源文件名",
        "篇名",
        "出版时间",
        "出版物名称",
        "品种/处理组",
        "地点/材料",
        "时间点",
        "证据状态",
    ]
    main = workbook["dataset_main"]
    assert [cell.value for cell in main[1]] == [*base, "温度"]
    data_row = [cell.value for cell in main[2]]
    assert data_row[0] == "paperA.pdf"
    assert data_row[1] == "论文A"
    assert data_row[2] == "2023"
    assert data_row[4] == "A"
    assert data_row[5] == "北京"
    assert data_row[6] == "48h"
    assert data_row[7] == "confirmed"
    assert data_row[8] == "25"

    ml = workbook["dataset_ml"]
    assert [cell.value for cell in ml[1]] == [*base, "温度"]
    assert [cell.value for cell in ml[2]][8] == 25

    assert [cell.value for cell in workbook["token_dictionary"][1]] == [
        "标准名",
        "类别",
        "别名",
        "首选单位",
        "状态",
    ]
    assert [cell.value for cell in workbook["token_dictionary"][2]] == [
        "温度",
        "工艺参数",
        "temp",
        "%",
        "confirmed",
    ]
    assert [cell.value for cell in workbook["traceability"][1]] == [
        "row_key",
        "field_key",
        "raw_value",
        "page_no",
        "evidence_text",
    ]
    assert [cell.value for cell in workbook["conversion_records"][1]] == [
        "原值",
        "原单位",
        "标准值",
        "标准单位",
        "规则/公式",
        "状态",
        "时间",
    ]
    conversion_row = [cell.value for cell in workbook["conversion_records"][2]]
    assert conversion_row[1] == "%"
    assert conversion_row[4] == "x+273.15"
    assert [cell.value for cell in workbook["audit_log"][1]] == [
        "时间",
        "操作",
        "对象类型",
        "对象ID",
        "执行人",
        "摘要",
    ]
    audit_row = [cell.value for cell in workbook["audit_log"][2]]
    assert audit_row[1] == "update"
    assert audit_row[4] == "张三"
    assert audit_row[5] == "人工修改"


def test_search_run_export_sheets_and_project_isolation(export_env) -> None:
    session, tables, storage = export_env
    project_id = uuid4()
    document_id, version_id, page_id = _insert_common_document(session, tables, project_id)

    run_id = uuid4()
    category_id = uuid4()
    term_id = uuid4()

    session.execute(
        tables["search_runs"].insert().values(id=run_id, project_id=project_id)
    )
    session.execute(
        tables["search_results"].insert().values(
            id=uuid4(),
            search_run_id=run_id,
            result_no=1,
            document_version_id=version_id,
            page_id=page_id,
            evidence_type="text",
            previous_context="上一句内容",
            matched_context="命中温度句",
            next_context="下一句内容",
            matched_terms=[
                {"term": "温度", "matched": True},
                {"term": "产率", "matched": False},
            ],
            review_status="pending",
        )
    )
    session.execute(
        tables["document_tables"].insert().values(
            id=uuid4(),
            document_version_id=version_id,
            page_id=page_id,
            table_no="表1",
            title="表标题",
            caption="表注",
            structured_data={"rows": [[1, 2], [3, 4]]},
        )
    )
    session.execute(
        tables["document_figures"].insert().values(
            id=uuid4(),
            document_version_id=version_id,
            page_id=page_id,
            figure_no="图1",
            title="图标题",
            caption="图注",
            axis_metadata={"x": "time"},
            legend_metadata={},
            extracted_labels=[],
            semantic_summary="图摘要",
        )
    )
    session.execute(
        tables["term_categories"].insert().values(id=category_id, name="工艺参数")
    )
    session.execute(
        tables["terms"].insert().values(
            id=term_id,
            project_id=project_id,
            category_id=category_id,
            canonical_name="温度",
            status="candidate",
            is_selected=False,
            preferred_unit_id=None,
            deleted_at=None,
        )
    )
    session.execute(
        tables["term_aliases"].insert().values(id=uuid4(), term_id=term_id, alias_text="temp")
    )
    session.execute(
        tables["term_occurrences"].insert().values(
            id=uuid4(),
            project_id=project_id,
            term_id=term_id,
            document_version_id=version_id,
            context_text="温度示例证据",
            occurrence_count=4,
        )
    )
    session.commit()

    service = SearchExportService(session, storage)
    path = service.export_xlsx(project_id, run_id)
    workbook = load_workbook(BytesIO(Path(path).read_bytes()))

    assert workbook.sheetnames == [
        "documents",
        "search_results",
        "figure_table_results",
        "token_candidates",
    ]
    assert [cell.value for cell in workbook["documents"][1]] == [
        "标题",
        "作者",
        "出版时间",
        "出版物",
        "品种/处理组",
        "地点/材料",
        "文件名",
    ]
    documents_row = [cell.value for cell in workbook["documents"][2]]
    assert documents_row[0] == "论文A"
    assert documents_row[1] == "李雷、韩梅梅"
    assert documents_row[6] == "paperA.pdf"

    assert [cell.value for cell in workbook["search_results"][1]] == [
        "序号",
        "文献ID",
        "标题",
        "检索词",
        "证据类型",
        "页码",
        "上一句",
        "命中句",
        "下一句",
        "合并证据内容",
        "审核状态",
    ]
    result_row = [cell.value for cell in workbook["search_results"][2]]
    assert result_row[3] == "温度"
    assert result_row[5] == 3
    assert result_row[9] == "上一句内容 命中温度句 下一句内容"

    assert [cell.value for cell in workbook["figure_table_results"][1]] == [
        "表号/图号",
        "表题/图题",
        "图注",
        "结构化内容",
        "页码",
        "所属文献",
    ]
    figure_table_values = {
        row[0].value for row in workbook["figure_table_results"].iter_rows(min_row=2)
    }
    assert {"表1", "图1"} <= figure_table_values

    assert [cell.value for cell in workbook["token_candidates"][1]] == [
        "候选词",
        "类别",
        "出现次数",
        "出现文献数",
        "别名建议",
        "示例证据",
        "人工选择状态",
    ]
    token_row = [cell.value for cell in workbook["token_candidates"][2]]
    assert token_row[0] == "温度"
    assert token_row[2] == 4
    assert token_row[3] == 1
    assert token_row[4] == "temp"
    assert token_row[5] == "温度示例证据"

    with pytest.raises(AppError):
        service.export_xlsx(uuid4(), run_id)
